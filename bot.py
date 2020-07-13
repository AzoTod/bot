from openpyxl import load_workbook
from telethon import TelegramClient, sync, events


api_id = 1660456                 
api_hash = "891e667b9f789a703506820bbf0c6154"  
client = TelegramClient('sesion_name', api_id, api_hash)
print('1')
     
@client.on(events.NewMessage(chats=('statamk10')))
async def normal_handler(event):
     message=(event.message.to_dict()['message'])
     last = message
     print('2')
     message=message.lower()
          
     wb = load_workbook('./table.xlsx')
     sheet = wb.worksheets[0]
     row_count = sheet.max_row
     print('3')
     for i in range(2, row_count+1):
               
          pers_left=sheet.cell(row=i, column= 1).value
               
          pers_left=pers_left.replace(" ", "")         
          pers_left=pers_left.replace("-", "")
          pers_right=sheet.cell(row=i, column=2).value
          pers_right=pers_right.replace(" ", "")
          pers_right=pers_right.replace("-", "")    
          obshee=(pers_left+pers_right).lower()
          print('4')     
               
          znach=pers_right=sheet.cell(row=i, column=3).value
          znach=znach.lower()
          
          if message.find(obshee) != -1 :
               if znach=='б':
                    ss = (last+'\n ФАТ 2,5 БОЛЬШЕ')
                    print(ss)
                    break
               if znach=='м':
                    ss = (last+'\n ФАТ 2,5 МЕНЬШЕ')
                    print(ss)  
                    break
               
     print('5')
     await client.send_message('https://t.me/joinchat/L9KUnBVGn7wxg4WO265gWw', ss)
     print('6')
client.start()

client.run_until_disconnected()


